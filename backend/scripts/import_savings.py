"""
Import women savings total.xlsx into PostgreSQL.

Usage (from backend/):
    source .venv/bin/activate
    python scripts/import_savings.py
"""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# Allow imports from backend/app
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.config import settings

XLSX_PATH = Path(__file__).parent.parent.parent / "docs" / "women savings total.xlsx"
TODAY = date.today()

# Canonical village name for each sheet tab
SHEET_TO_VILLAGE: dict[str, str] = {
    "sittilingi": "Sittilingi",
    "moolssit": "Moola Sittilingi",
    "Korayar": "Korayar",
    "Thekkanampatti": "Thekkanampatti",
    "S Dadampatti": "S Dadampatti",
    "Mullikadu": "Mullikadu",
    "Erattakuttai": "Erattakuttai",
    "Mel Thanda": "Mel Thanda",
    "AK Thanda": "AK Thanda",
    "palakuittai": "Palakuttai",
    "naikuthi": "Naikuthi",
    "vannadurai": "Vannadurai",
    "malaithangi": "Malaithangi",
    "kalianku": "Kaliankuttai",
    "nammangad": "Nammangadu",
    "velanur": "Velanur",
    "kathripatti": "Kathripatti",
}


def to_float(val) -> float:
    """Return a float for numeric cell values; 0.0 for None/non-numeric."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return 0.0


def parse_sheet(ws, village_name: str) -> list[dict]:
    """
    Parse one village sheet into a list of group-month records.

    Each record:
        group_name, village_name,
        entry_month (date),
        savings, int_principal, int_interest, to_bank, from_bank,
        sofa1_disbursed, sofa1_repayment, sofa1_interest,
        sofa2_disbursed, sofa2_repayment, sofa2_interest
    """
    rows = list(ws.iter_rows(values_only=True))
    group_starts = [i for i, row in enumerate(rows) if row[0] == "Gr. Name"]
    records = []

    for g in group_starts:
        group_name = rows[g][1]
        # Skip placeholder rows with no group name
        if not group_name or not str(group_name).strip():
            continue

        group_name = str(group_name).strip()

        # Each block has four date rows at offsets +3, +8, +13, +18
        # All four date rows are identical — use offset +3
        date_row = rows[g + 3] if g + 3 < len(rows) else []

        # Collect all valid (col_index, meeting_date) pairs from col C (index 2) onwards
        dated_cols: list[tuple[int, date]] = []
        for col_idx, cell in enumerate(date_row):
            if col_idx < 2:
                continue
            if not isinstance(cell, datetime):
                continue
            meeting_date = cell.date()
            if meeting_date > TODAY:
                continue
            dated_cols.append((col_idx, meeting_date))

        if not dated_cols:
            continue

        # Metric rows at fixed offsets from group header
        savings_row    = rows[g + 4]  if g + 4  < len(rows) else []
        int_prin_row   = rows[g + 5]  if g + 5  < len(rows) else []
        int_int_row    = rows[g + 6]  if g + 6  < len(rows) else []
        # Opening bank balance: find the row labelled "Opening Balance" within the bank
        # section (offsets +8 to +12) and take the first numeric value in that row.
        # The value column varies by group, so we search by label rather than fixed index.
        opening_balance = 0.0
        for _off in range(g + 8, min(g + 13, len(rows))):
            _row = rows[_off]
            label = next((c for c in _row[:3] if isinstance(c, str)), "")
            if "opening balance" in label.lower():
                for _col in range(2, len(_row)):
                    _v = _row[_col]
                    if isinstance(_v, (int, float)) and _v != 0:
                        opening_balance = float(_v)
                        break
                break
        to_bank_row    = rows[g + 10] if g + 10 < len(rows) else []
        from_bank_row  = rows[g + 11] if g + 11 < len(rows) else []
        sofa1_dis_row  = rows[g + 14] if g + 14 < len(rows) else []
        sofa1_ret_row  = rows[g + 15] if g + 15 < len(rows) else []
        sofa1_int_row  = rows[g + 16] if g + 16 < len(rows) else []
        sofa2_dis_row  = rows[g + 19] if g + 19 < len(rows) else []
        sofa2_ret_row  = rows[g + 20] if g + 20 < len(rows) else []
        sofa2_int_row  = rows[g + 21] if g + 21 < len(rows) else []

        def get(row, col_idx):
            val = row[col_idx] if col_idx < len(row) else None
            # Guard against date objects leaked into numeric cells
            if isinstance(val, (datetime, date)):
                return 0.0
            return to_float(val)

        for col_idx, meeting_date in dated_cols:
            savings         = get(savings_row,   col_idx)
            int_principal   = get(int_prin_row,  col_idx)
            int_interest    = get(int_int_row,   col_idx)
            to_bank         = get(to_bank_row,   col_idx)
            from_bank       = get(from_bank_row, col_idx)
            sofa1_dis       = get(sofa1_dis_row, col_idx)
            sofa1_ret       = get(sofa1_ret_row, col_idx)
            sofa1_int       = get(sofa1_int_row, col_idx)
            sofa2_dis       = get(sofa2_dis_row, col_idx)
            sofa2_ret       = get(sofa2_ret_row, col_idx)
            sofa2_int       = get(sofa2_int_row, col_idx)

            # Skip months where everything is zero
            all_values = [savings, int_principal, int_interest, to_bank, from_bank,
                          sofa1_dis, sofa1_ret, sofa1_int, sofa2_dis, sofa2_ret, sofa2_int]
            if not any(v != 0.0 for v in all_values):
                continue

            records.append({
                "group_name": group_name,
                "village_name": village_name,
                "entry_month": meeting_date,
                "opening_balance": opening_balance,
                "savings": savings,
                "int_principal": int_principal,
                "int_interest": int_interest,
                "to_bank": to_bank,
                "from_bank": from_bank,
                "sofa1_dis": sofa1_dis,
                "sofa1_ret": sofa1_ret,
                "sofa1_int": sofa1_int,
                "sofa2_dis": sofa2_dis,
                "sofa2_ret": sofa2_ret,
                "sofa2_int": sofa2_int,
            })

    return records


def run():
    engine = create_engine(settings.database_url, future=True)

    print(f"Loading {XLSX_PATH.name} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_records: list[dict] = []
    for sheet_name, village_name in SHEET_TO_VILLAGE.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        records = parse_sheet(ws, village_name)
        print(f"  {sheet_name:22s} -> {village_name:22s}  {len(records):4d} month-records")
        all_records.extend(records)

    print(f"\nTotal month-records to insert: {len(all_records)}")

    with Session(engine) as session:
        # --- groups ---
        # Take opening_balance from the first record seen for each group
        group_opening: dict[tuple[str, str], float] = {}
        for r in all_records:
            key = (r["group_name"], r["village_name"])
            if key not in group_opening:
                group_opening[key] = r["opening_balance"]

        group_keys = {(r["group_name"], r["village_name"]) for r in all_records}
        group_id: dict[tuple[str, str], int] = {}
        for group_name, vill_name in sorted(group_keys):
            # Code includes village to guarantee global uniqueness (max 50 chars)
            code = f"{group_name} ({vill_name})"[:50]
            result = session.execute(
                text("""
                    INSERT INTO groups
                        (name, village_name, code, register_template, is_active,
                         opening_bank_balance, created_at, updated_at)
                    VALUES
                        (:name, :village_name, :code, 'default_v1', true,
                         :opening_balance, NOW(), NOW())
                    ON CONFLICT (code) DO UPDATE SET
                        name = EXCLUDED.name,
                        opening_bank_balance = EXCLUDED.opening_bank_balance
                    RETURNING id
                """),
                {
                    "name": group_name,
                    "village_name": vill_name,
                    "code": code,
                    "opening_balance": group_opening[(group_name, vill_name)],
                },
            )
            group_id[(group_name, vill_name)] = result.scalar_one()
        print(f"Inserted {len(group_id)} groups")

        # --- month_entries + sofa_loan_entries ---
        entry_count = 0
        for r in all_records:
            gid = group_id[(r["group_name"], r["village_name"])]
            result = session.execute(
                text("""
                    INSERT INTO month_entries (
                        group_id, entry_month, entry_mode, status,
                        savings_collected, internal_loan_principal_disbursed,
                        internal_loan_interest_collected, to_bank, from_bank,
                        sofa_loan_disbursed, sofa_loan_repayment, sofa_loan_interest_collected,
                        warning_flags, source_count, created_at, updated_at
                    ) VALUES (
                        :group_id, :entry_month, 'manual', 'synced',
                        :savings, :int_principal, :int_interest, :to_bank, :from_bank,
                        :sofa_dis, :sofa_ret, :sofa_int,
                        '[]', 0, NOW(), NOW()
                    )
                    ON CONFLICT (group_id, entry_month) DO UPDATE SET
                        savings_collected               = EXCLUDED.savings_collected,
                        internal_loan_principal_disbursed = EXCLUDED.internal_loan_principal_disbursed,
                        internal_loan_interest_collected = EXCLUDED.internal_loan_interest_collected,
                        to_bank                         = EXCLUDED.to_bank,
                        from_bank                       = EXCLUDED.from_bank,
                        sofa_loan_disbursed             = EXCLUDED.sofa_loan_disbursed,
                        sofa_loan_repayment             = EXCLUDED.sofa_loan_repayment,
                        sofa_loan_interest_collected    = EXCLUDED.sofa_loan_interest_collected,
                        updated_at                      = NOW()
                    RETURNING id
                """),
                {
                    "group_id": gid,
                    "entry_month": r["entry_month"],
                    "savings": r["savings"],
                    "int_principal": r["int_principal"],
                    "int_interest": r["int_interest"],
                    "to_bank": r["to_bank"],
                    "from_bank": r["from_bank"],
                    "sofa_dis": r["sofa1_dis"] + r["sofa2_dis"],
                    "sofa_ret": r["sofa1_ret"] + r["sofa2_ret"],
                    "sofa_int": r["sofa1_int"] + r["sofa2_int"],
                },
            )
            entry_id = result.scalar_one()
            entry_count += 1

            # Populate per-slot breakdown in sofa_loan_entries
            for slot, (dis, ret, intr) in enumerate(
                [
                    (r["sofa1_dis"], r["sofa1_ret"], r["sofa1_int"]),
                    (r["sofa2_dis"], r["sofa2_ret"], r["sofa2_int"]),
                ],
                start=1,
            ):
                if dis or ret or intr:
                    session.execute(
                        text("""
                            INSERT INTO sofa_loan_entries
                                (month_entry_id, loan_slot, disbursed, repayment,
                                 interest_collected, created_at, updated_at)
                            VALUES
                                (:entry_id, :slot, :disbursed, :repayment, :interest, NOW(), NOW())
                            ON CONFLICT (month_entry_id, loan_slot) DO UPDATE SET
                                disbursed          = EXCLUDED.disbursed,
                                repayment          = EXCLUDED.repayment,
                                interest_collected = EXCLUDED.interest_collected,
                                updated_at         = NOW()
                        """),
                        {
                            "entry_id": entry_id,
                            "slot": slot,
                            "disbursed": dis,
                            "repayment": ret,
                            "interest": intr,
                        },
                    )

        session.commit()

    print(f"Inserted {entry_count} month_entries")
    print("Done.")


if __name__ == "__main__":
    run()
